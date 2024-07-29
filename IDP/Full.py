import os
import logging
import requests
import cv2
import pytesseract
import time
import threading
import datetime
import openpyxl
import json
import pandas as pd
import sys
from PyQt5 import QtWidgets,QtCore,QtGui
from PyQt5.QtWidgets import QLabel, QInputDialog, QApplication , QMainWindow, QMessageBox, QLineEdit,QAction,QInputDialog
from PyQt5.QtGui import QDesktopServices,QIcon
from PyQt5.QtCore import QUrl

logging.basicConfig(filename="mylogs.log", level=logging.DEBUG, format='%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
logging.debug("Debug message")
logging.info("Info message")
logging.warning("Warning message")
logging.error("Error message")
logging.critical("Critical message")

class Choice:
    def __init__(self, input_str):
        self.input_str = input_str
        self.counts = {"a": 0, "b": 0, "c": 0}
        self.api_key = 'T0EMLF3D148AXJKJ'
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)

    def update_api_key(self, new_key):
        self.api_key = new_key

    def _validate_api_key(self):
        if not self.api_key or not isinstance(self.api_key, str):
            self.logger.error("Invalid API key")
            raise ValueError("Invalid API key")

    def _validate_field_numbers(self, field_name):
        field_number_map = {"a": 1, "b": 2, "c": 3}
        if self.input_str not in field_number_map:
            self.logger.error(f"Invalid field name: {self.input_str}")
            raise ValueError(f"Invalid field name: {self.input_str}")
        return field_number_map[self.input_str]

    def update_field(self, field_name):
        self._validate_api_key()
        field_number = self._validate_field_numbers(field_name)
        if self.counts[self.input_str] == 0:
            self.counts[self.input_str] += 1
            endpoint = f'https://api.thingspeak.com/update?api_key={self.api_key}&field{field_number}={self.counts[self.input_str]}'
            try:
                response = requests.get(endpoint)
                response.raise_for_status()
                self.logger.info(f"Field {field_number} is updated")
            except requests.exceptions.HTTPError as errh:
                self.logger.error(f"Http Error updating field {field_number} during first sending: {errh}")
            except requests.exceptions.ConnectionError as errc:
                self.logger.error(f"Error Connecting updating field {field_number} during first sending: {errc}")
            except requests.exceptions.Timeout as errt:
                self.logger.error(f"Timeout Error updating field {field_number} during first sending: {errt}")
            except requests.exceptions.RequestException as err:
                self.logger.error(f"Something went wrong updating field {field_number} during first sending: {err}")
            time.sleep(15)
            self.send_zero_update(field_name)

    def send_zero_update(self, field_name):
        if self.counts[self.input_str] == 1:
            self.counts[self.input_str] = 0
            endpoint = f'https://api.thingspeak.com/update?api_key={self.api_key}&field{field_name}={self.counts[self.input_str]}'
            try:
                response = requests.get(endpoint)
                response.raise_for_status()
                logging.info(f"Field{field_name} is reset")
            except requests.exceptions.HTTPError as errh:
                logging.error("Http Error: %s", errh)
            except requests.exceptions.ConnectionError as errc:
                logging.error("Error Connecting: %s", errc)
            except requests.exceptions.Timeout as errt:
                logging.error("Timeout Error: %s", errt)
            except requests.exceptions.RequestException as err:
                logging.error("Something went wrong: %s", err)
            response = requests.get(endpoint)

    def print_output(self):
        if self.input_str in self.counts:
            logging.info(f"it is {self.input_str}")
        if self.input_str == "a":
            self.update_field(1) 
        elif self.input_str == "b":
            self.update_field(2)
        elif self.input_str == "c":
            self.update_field(3) 
        else:
            logging.warning("not a,b or c")

class OCR:
    def __init__(self):
        self.is_running = False
        self.path_to_tesseract = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        self.send_thread_running = threading.Event()
        logging.basicConfig(filename='OCR.log', level=logging.ERROR, 
                            format='%(asctime)s %(levelname)s %(message)s')
        try:
            self.camera = cv2.VideoCapture(0)
            if not self.camera.isOpened():
                raise Exception("Error: Unable to open camera")
        except Exception as e:
            logging.error(f"{e}")
            self.camera.release()
            exit()

    def run(self):
        print("Started OCR")
        self.is_running = True
        while True:
            ret, frame = self.camera.read()
            if not ret:
                logging.error("Error: unable to capture from camera")
                self.camera.release()
                break
     
            frame = cv2.medianBlur(frame, 7)
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 8)
            #thresh = cv2.bitwise_not(thresh)
            #kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (20,20 ))
            #thresh = cv2.dilate(thresh, kernel, iterations = 1)
            #thresh = cv2.bitwise_not(thresh)

            if thresh is None or cv2.countNonZero(thresh) == 0:
                logging.warning("Error: Image is all black or empty")
                continue
            try:
                text = pytesseract.image_to_string(thresh, lang='eng', config='--psm 7')
            except Exception as e:
                logging.error(f"Error: Unable to read text with Tesseract: {e}")
                continue
            text = text.lower()
            
            cv2.putText(thresh, text, (10,50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255), 2, cv2.LINE_AA)
            cv2.imshow('frame', thresh)

            if not self.send_thread_running.is_set():
                self.send_thread_running.set()
                threading.Thread(target=self.send, args=(text,)).start()
            key = cv2.waitKey(1)
            if key == ord('s') or self.is_running == False:  
                break
        self.camera.release()
        cv2.destroyAllWindows()

    def send(self, text):
        for char in text:
            if char in ["a","b","c"]:
                
                assign = Assign(char)
                assign.Timing(char)
                choice = Choice(char)
                choice.print_output()
                time.sleep(15)
                break
        self.send_thread_running.clear()
    
    def stop(self):
        print("Stopped OCR")
        self.is_running = False
        self.camera.release()

class Assign:
    def __init__(self, choice):
        try:
            with open("counts_ocr.json", "r") as f:
                self.counts_ocr = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.counts_ocr = {"a": 0, "b": 0, "c": 0}

        try:
            with open("lots_ocr.json", "r") as f:
                self.lots_ocr = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.lots_ocr = {"a": 0, "b": 0, "c": 0}

        self.choice = choice
        self.start_time = None
        self.end_time = None
        self.lot = f"{choice.upper()}000"
        
    def Timing(self, choice):
        if self.counts_ocr[choice] == 0:
            self.start_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        self.counts_ocr[choice] += 1
        with open("counts_ocr.json", "w") as f:
            json.dump(self.counts_ocr, f)

        if self.counts_ocr[choice] == 5:
            self.end_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"Start time for {choice}: {self.start_time}")
            print(f"End time for {choice}: {self.end_time}")
            self.counts_ocr[choice] = 0
            self.lots_ocr[choice] += 1
            formatted_lot = self.Lot(choice)
            self.write_to_excel(choice, self.start_time, self.end_time, formatted_lot)
        else:
            formatted_lot = self.Lot(choice)
            self.write_to_excel(choice, self.start_time, None, formatted_lot)
        with open("counts_ocr.json", "w") as f:
            json.dump(self.counts_ocr, f)
        with open("lots_ocr.json", "w") as f:
            json.dump(self.lots_ocr, f)
        #Cnt = MyWindow()
        #Cnt.update_count(choice)
    
    def Lot(self, choice):
        self.lot = self.lot[:-3] + str(int(self.lot[-3:]) + self.lots_ocr[choice]).zfill(3)
        if self.lot[-3:] == '999':
            self.lots_ocr[choice] = 0
            self.lot = self.lot[:-3] + '000'
            self.lot = chr(ord(self.lot[0]) + 1) + self.lot[1:]
        return self.lot

    def write_to_excel(self, choice, start_time, end_time, lot):
        MAX_SHEETS = 65535
        choice = choice.upper()

        if os.path.isfile("data.xlsx"):
            try:
                wb = openpyxl.load_workbook("data.xlsx")
            except PermissionError:
                print("The file is open by another process, please close it and try again.")
                return
        else:
            wb = openpyxl.Workbook()

        if choice not in wb.sheetnames:
            ws = wb.create_sheet(choice)
            ws.append(["Choice", "Start Time", "End Time", "Lot"])
        else:
            ws = wb[choice]

        if len(wb.sheetnames) >= MAX_SHEETS:
            print("The maximum number of worksheets in the Excel file has been reached. Please remove some worksheets before adding new ones.")
            return

        if ws.max_row >= 1048576:
            new_sheet_name = f"{choice}({ws.max_row//1048576})"
            ws = wb.create_sheet(new_sheet_name)
            ws.append(["Choice", "Start Time", "End Time", "Lot"])
        ws.append([choice, start_time, end_time, lot])

        try:
            wb.save("data.xlsx")
        except openpyxl.utils.exceptions.InvalidFileException as e:
            print(f"Error while saving the file: {e}")
            return

        try:
            with open("counts_ocr.json", "w") as f:
                json.dump(self.counts_ocr, f)
        except:
            print("Error while saving the counts")

class MyWindow(QtWidgets.QMainWindow):
    def __init__(self,):
        super().__init__()
        self.choice = Choice("w")
        try:
            with open("counts_ocr.json", "r") as f:
                self.counts_ocr = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.counts_ocr = {"a": 0, "b": 0, "c": 0}
        
        self.my_class = OCR()

        self.setWindowTitle("My GUI")
        self.setGeometry(200, 200, 400, 400)
        self.setAutoFillBackground(True)
        palette = self.palette()
        self.setStyleSheet("background-color:  #464D46;")
        self.setPalette(palette)

        central_widget = QtWidgets.QWidget(self)
        self.setCentralWidget(central_widget)

        self.start_button = QtWidgets.QPushButton("Start", self)
        self.start_button.setAutoFillBackground(True)
        self.start_button.setStyleSheet("background-color:  #8d9db6;")
        self.start_button.move(20, 70)
        self.start_button.clicked.connect(self.start_my_class)

        self.stop_button = QtWidgets.QPushButton("Stop", self)
        self.stop_button.setAutoFillBackground(True)
        self.stop_button.setStyleSheet("background-color:  #667292;")
        self.stop_button.move(120, 70)
        self.stop_button.clicked.connect(self.stop_my_class)
        
        self.A_label = QLabel(self)
        self.A_label.setText("A : " + str(self.counts_ocr["a"]))
        self.A_label.move(350, 300)

        self.B_label = QLabel(self)
        self.B_label.setText("B : " + str(self.counts_ocr["b"]))
        self.B_label.move(350, 325)

        self.C_label = QLabel(self)
        self.C_label.setText("C : " + str(self.counts_ocr["c"]))
        self.C_label.move(350, 350)

        self.api_key_input = QtWidgets.QLineEdit(self)
        self.api_key_input.setPlaceholderText("Enter new API key here")
        self.api_key_input.move(20, 110)

        self.updateButton = QtWidgets.QPushButton("Update API Key",self)
        self.updateButton.setAutoFillBackground(True)
        self.updateButton.setStyleSheet("background-color:  #8d9db6;")
        self.updateButton.move(120, 110)
        self.updateButton.clicked.connect(self.update_api_key)

        self.update_channel_action = QAction(QIcon("C:/Users/Kygo/Py_Projects/icon.jpeg"), "Update Channel ID", self)
        self.update_channel_action.triggered.connect(self.update_channel)
        self.toolbar = self.addToolBar("Update Channel")
        self.toolbar.addAction(self.update_channel_action)
        self.toolbar.setStyleSheet("background-color:  #4D4646;")
        self.channel_id = "2005241"

        self.toolbar = self.addToolBar("My Toolbar")
        self.statusbar = self.statusBar()

        self.action1 = QtWidgets.QAction("View", self)
        self.action1.triggered.connect(self.View)

        self.action2 = QtWidgets.QAction("Thresh", self)
        self.action2.triggered.connect(self.Thresh)
        
        self.action3 = QtWidgets.QAction("Excel", self)
        self.action3.triggered.connect(self.open_excel_file)
        self.menu_bar = self.menuBar()
        self.file_menu = self.menu_bar.addMenu("File")
        self.file_menu.addAction(self.action3)

        self.action4 = QtWidgets.QAction("Channel", self)
        self.action4.triggered.connect(self.Channel)

        self.action5 = QtWidgets.QAction("Detect", self)
        self.action5.triggered.connect(self.Detect)

        self.checkWifiButton = QtWidgets.QToolButton(self)
        self.checkWifiButton.setText("Check Wifi")
        self.checkWifiButton.setIcon(QtGui.QIcon("C:\\Users\\Kygo\\Py_Projects\\wifi.jpeg"))
        self.checkWifiButton.clicked.connect(self.checkWifi)

        self.toolbar.addAction(self.action1)
        self.toolbar.setStyleSheet("background-color:  #9B8C8C;")
        self.toolbar.addSeparator()
        self.toolbar.addAction(self.action2)
        self.toolbar.addSeparator()
        self.toolbar.addAction(self.action3)
        self.toolbar.addSeparator()
        self.toolbar.addAction(self.action4)
        self.toolbar.addSeparator()
        self.toolbar = self.addToolBar("Tools")
        self.toolbar.addWidget(self.checkWifiButton)
        self.toolbar.setStyleSheet("background-color:  #4D4646;")
        self.toolbar.addSeparator()
        self.toolbar.addAction(self.action5)

        self.date_label = QLabel(self)
        self.date_label.setText("")
        self.date_label.move(30, 330)
        
        self.time_label = QLabel(self)
        self.time_label.setText("")
        self.time_label.move(30, 350)
        
        self.timer = QtCore.QTimer(self)
        self.timer.timeout.connect(self.show_time)
        self.timer.start(1000)

        self.show()
    
    def checkWifi(self):
        output = os.popen("iwconfig wlan0").read()
        if "Link Quality=" in output:
            wifi_strength = output.split("Link Quality=")[1].split("  ")[0].split("/")[0]
            return int(wifi_strength)
        else:
            return -1

    def show_time(self):
        current_time = datetime.datetime.now().strftime("%H:%M:%S")
        current_date = datetime.datetime.now().strftime("%Y-%m-%d")
        self.time_label.setText(current_time)
        self.date_label.setText(current_date)

    def Channel(self):
        url = QUrl("https://thingspeak.com/channels/2005241")
        QDesktopServices.openUrl(url)
    
    def update_channel(self):
        text, ok = QInputDialog.getText(self, "Update Channel ID", "Enter new channel ID:")
        if ok:
            self.channel_id = text

    def open_excel_file(self):
        file_name, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)")
        if file_name:
            os.startfile(file_name)

    def update_api_key(self):
        text, ok = QInputDialog.getText(self, "Update API Key", "Enter new API key:", QLineEdit.Normal, "")
        if ok:
            new_key = text
            if len(new_key) <= 16 and new_key.isalnum():
                self.updateButton.setEnabled(True)
                self.choice.update_api_key(new_key)
                self.updateButton.setEnabled(False)
            else:
                QMessageBox.warning(self, "Warning", "API key must be 16 characters or less and alphanumeric.")
    
    def View(self):
        try:
            self.camera = cv2.VideoCapture(0)
            if not self.camera.isOpened():
                raise Exception("Error: Unable to open camera")
        except Exception as e:
            logging.error(f"{e}")
            self.camera.release()
            exit()

        while True:
            ret, frame = self.camera.read()
            if ret:
                cv2.imshow("Camera", frame)
                key = cv2.waitKey(1)
                if key == ord('w'):
                    self.camera.release()
                    self.action1.setEnabled(True)
                    cv2.destroyAllWindows()
                    break
            else:
                logging.error("Error: unable to capture from camera")
                self.camera.release()
                self.action1.setEnabled(True)
                break
    
    def Thresh(self):
        try:
            self.camera = cv2.VideoCapture(0)
            if not self.camera.isOpened():
                raise Exception("Error: Unable to open camera")
        except Exception as e:
            logging.error(f"{e}")
            self.camera.release()
            exit()

        while True:
            ret, frame = self.camera.read()
            frame = cv2.medianBlur(frame, 7)
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 8)

            if ret:
                cv2.imshow("Camera", thresh)
                key = cv2.waitKey(1)
                if key == ord('w'):
                    self.camera.release()
                    self.action2.setEnabled(True)
                    cv2.destroyAllWindows()
                    break
            else:
                logging.error("Error: unable to capture from camera")
                self.camera.release()
                self.action2.setEnabled(True)
                break
    
    def Detect(self):
        try:
            self.camera = cv2.VideoCapture(0)
            if not self.camera.isOpened():
                raise Exception("Error: Unable to open camera")
        except Exception as e:
            logging.error(f"{e}")
            self.camera.release()
            exit()

        while True:
            ret, frame = self.camera.read()
            frame = cv2.medianBlur(frame, 7)
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 8)
            if thresh is None or cv2.countNonZero(thresh) == 0:
                logging.warning("Error: Image is all black or empty")
                continue
            try:
                text = pytesseract.image_to_string(thresh, lang='eng', config='--psm 7')
            except Exception as e:
                logging.error(f"Error: Unable to read text with Tesseract: {e}")
                continue
            text = text.lower()
                
            cv2.putText(thresh, text, (10,50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255), 2, cv2.LINE_AA)  
            if ret:
                cv2.imshow("Camera", thresh)
                key = cv2.waitKey(1)
                if key == ord('w'):
                    self.camera.release()
                    self.action2.setEnabled(True)
                    cv2.destroyAllWindows()
                    break
            else:
                logging.error("Error: unable to capture from camera")
                self.camera.release()
                self.action2.setEnabled(True)
                break    

    def start_my_class(self):
        if not self.my_class.is_running:
            self.my_class.run()
            self.start_button.setEnabled(False)
            self.stop_button.setEnabled(True)

    def stop_my_class(self):
        if self.my_class.is_running:
            self.my_class.stop()
            self.start_button.setEnabled(True)
            self.stop_button.setEnabled(False)

    def update_count(self, choice):
        try:
            with open("counts_ocr.json", "r") as f:
                self.counts_ocr = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            self.counts_ocr = {"a": 0, "b": 0, "c": 0}
        
        self.counts_ocr[choice] += 1
            
        self.A_label.setText(str(self.counts_ocr['a']))
        self.B_label.setText(str(self.counts_ocr['b']))
        self.C_label.setText(str(self.counts_ocr['c']))

        self.write_to_screen(choice)

    def write_to_screen(self, choice):
        if choice == "a":
            self.A_label.setText("A: " + str(self.counts_ocr["a"]))
        elif choice == "b":
            self.B_label.setText("B: " + str(self.counts_ocr["b"]))
        elif choice == "c":
            self.C_label.setText("C: " + str(self.counts_ocr["c"]))

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MyWindow()
    sys.exit(app.exec_())
